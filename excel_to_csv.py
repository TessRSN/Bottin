#!/usr/bin/env python3
"""
Convertit le fichier Excel RSN en CSV en préservant les hyperliens.

Pour les colonnes qui contiennent des hyperliens (ORCID, CV/LinkedIn, E-mail...),
le script remplace le texte affiché par l'URL réelle du lien.

Usage:
    python excel_to_csv.py RSN_BD_AllMembers.xlsx all_members.csv
    python excel_to_csv.py RSN_BD_AllMembers.xlsx  # -> all_members.csv par défaut
    python excel_to_csv.py RSN_BD_AllMembers.xlsx all_members.csv "Ma feuille"
"""

import sys
import csv
import openpyxl


def excel_to_csv(input_xlsx, output_csv="all_members.csv", sheet_name="ALL (new)"):
    wb = openpyxl.load_workbook(input_xlsx)
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        print(f"⚠ Feuille '{sheet_name}' introuvable. Feuilles disponibles: {wb.sheetnames}")
        print(f"  → Utilisation de la première feuille: '{wb.worksheets[0].title}'")
        ws = wb.worksheets[0]

    rows = []
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        csv_row = []
        for cell in row:
            value = cell.value if cell.value is not None else ""
            # Si la cellule a un hyperlien, on utilise l'URL à la place du texte
            if cell.hyperlink and cell.hyperlink.target:
                value = cell.hyperlink.target
            csv_row.append(str(value).strip())
        rows.append(csv_row)

    with open(output_csv, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerows(rows)

    print(f"✓ {ws.max_row} lignes exportées vers {output_csv}")
    print(f"  (feuille: '{ws.title}', colonnes: {ws.max_column})")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python excel_to_csv.py <fichier.xlsx> [sortie.csv]")
        sys.exit(1)

    input_file = sys.argv[1]
    output_file = sys.argv[2] if len(sys.argv) > 2 else "all_members.csv"
    sheet = sys.argv[3] if len(sys.argv) > 3 else "ALL (new)"
    excel_to_csv(input_file, output_file, sheet)
